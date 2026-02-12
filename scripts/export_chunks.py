#!/usr/bin/env python3
"""Export document chunks and Ollama-generated questions to Excel.

Creates one Excel file per document in a dated subfolder, each containing:
- Chunks tab: all chunks for the document
- Questions tab: Ollama-recommended questions for the document

Usage:
    python scripts/export_chunks.py                   # output to data/chunks_MM_DD_YYYY/
    python scripts/export_chunks.py -o custom_dir     # custom output directory
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from qdrant_client import QdrantClient
from qdrant_client.models import FieldCondition, Filter, MatchValue

# Add project root to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from ai_ready_rag.config import get_settings  # noqa: E402
from ai_ready_rag.db.database import SessionLocal, init_db  # noqa: E402
from ai_ready_rag.db.models import Document  # noqa: E402

# Regex to strip characters illegal in Excel XML (control chars except tab/newline/cr)
_ILLEGAL_XML_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]")


def _sanitize(text: str) -> str:
    """Remove characters that openpyxl/Excel cannot store in a cell."""
    return _ILLEGAL_XML_CHARS.sub("", text) if text else text


# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
BOLD_FONT = Font(bold=True)


def get_ready_documents() -> list[dict]:
    """Fetch all documents with status=ready from the database."""
    init_db()
    db = SessionLocal()
    try:
        docs = (
            db.query(Document)
            .filter(Document.status == "ready")
            .order_by(Document.uploaded_at.desc())
            .all()
        )
        return [
            {
                "id": doc.id,
                "filename": doc.original_filename or doc.filename,
                "file_type": doc.file_type,
                "file_size": doc.file_size,
                "status": doc.status,
                "chunk_count": doc.chunk_count,
                "word_count": doc.word_count,
                "tags": ", ".join(t.name for t in doc.tags) if doc.tags else "",
                "uploaded_at": str(doc.uploaded_at) if doc.uploaded_at else "",
            }
            for doc in docs
        ]
    finally:
        db.close()


def get_chunks(qdrant: QdrantClient, document_id: str) -> list[dict]:
    """Fetch all chunks for a document from Qdrant."""
    points, _ = qdrant.scroll(
        collection_name="documents",
        scroll_filter=Filter(
            must=[FieldCondition(key="document_id", match=MatchValue(value=document_id))]
        ),
        with_payload=True,
        with_vectors=False,
        limit=1000,
    )

    chunks = []
    for p in points:
        payload = p.payload or {}
        chunks.append(
            {
                "index": payload.get("chunk_index", 0),
                "text": payload.get("chunk_text", ""),
                "words": len(payload.get("chunk_text", "").split()),
                "chars": len(payload.get("chunk_text", "")),
                "page": payload.get("page_number"),
                "section": payload.get("section"),
            }
        )

    return sorted(chunks, key=lambda c: c["index"])


def generate_questions(chunks: list[dict], doc_name: str, settings) -> list[str]:
    """Ask Ollama for recommended questions. Returns list of question strings."""
    if not chunks:
        return []

    sample_size = min(12, len(chunks))
    step = max(1, len(chunks) // sample_size)
    sampled = [chunks[i] for i in range(0, len(chunks), step)][:sample_size]

    total_words = sum(c["words"] for c in chunks)
    content_summary = "\n\n".join(f"[Chunk {c['index']}]: {c['text'][:300]}" for c in sampled)

    # Scale question count by document size
    if len(chunks) <= 3:
        num_questions = 3
    elif len(chunks) <= 10:
        num_questions = 5
    elif len(chunks) <= 50:
        num_questions = 10
    else:
        num_questions = 15

    prompt = f"""You are analyzing a document that has been chunked for a RAG system.
Document: {doc_name}
Total chunks: {len(chunks)}
Total words: {total_words:,}

Here are representative samples from the document:

{content_summary}

Based on this content, suggest exactly {num_questions} specific questions that a user could ask this RAG system.
Focus on questions that:
1. Can be answered from the document content
2. Cover different sections/topics in the document
3. Range from simple factual to analytical
4. Would be useful for someone who needs information from this document

Format: Number each question on its own line (1. question text). No other text."""

    try:
        with httpx.Client(timeout=300) as client:
            response = client.post(
                f"{settings.ollama_base_url}/api/generate",
                json={
                    "model": settings.chat_model,
                    "prompt": prompt,
                    "stream": False,
                    "options": {"temperature": 0.3},
                },
            )
            response.raise_for_status()
            answer = response.json().get("response", "").strip()

            # Parse numbered questions
            questions = []
            for line in answer.split("\n"):
                line = line.strip()
                if line and line[0].isdigit():
                    q = line.lstrip("0123456789").lstrip(".)- ").strip()
                    if q:
                        questions.append(q)
            return questions

    except httpx.TimeoutException:
        print(f"    WARNING: Ollama timed out for '{doc_name}'")
        return [f"[Ollama timed out — try: ollama pull {settings.chat_model}]"]
    except httpx.ConnectError:
        print(f"    WARNING: Cannot connect to Ollama at {settings.ollama_base_url}")
        return ["[Ollama not reachable]"]
    except Exception as e:
        print(f"    WARNING: Ollama error for '{doc_name}': {e}")
        return [f"[Error: {e}]"]


def styled_header(ws, headers: list[str], row: int = 1) -> None:
    """Write a styled header row."""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _safe_filename(name: str) -> str:
    """Convert a document name into a filesystem-safe filename (no extension)."""
    stem = Path(name).stem
    # Replace characters invalid on Windows/Linux
    safe = re.sub(r'[\\/*?:"<>|]', "_", stem)
    # Collapse runs of whitespace/underscores
    safe = re.sub(r"[\s_]+", "_", safe).strip("_")
    return safe[:120]  # Keep reasonable length


def main():
    parser = argparse.ArgumentParser(description="Export per-document chunk + question Excel files")
    parser.add_argument(
        "-o",
        "--output",
        help="Output directory (default: data/chunks_MM_DD_YYYY/)",
    )
    args = parser.parse_args()

    settings = get_settings()
    today = datetime.now().strftime("%m_%d_%Y")

    # Determine output directory
    out_dir = Path(args.output) if args.output else Path("data") / f"chunks_{today}"

    out_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("RAG Document & Chunk Export (per-file)")
    print(f"Output: {out_dir}/")
    print("=" * 60)

    # Step 1: Get documents
    print("\n[1/2] Fetching documents...")
    docs = get_ready_documents()
    print(f"  Found {len(docs)} ready documents")

    if not docs:
        print("No ready documents to export.")
        return

    # Step 2: One workbook per document
    print(f"\n[2/2] Exporting documents (model: {settings.chat_model})...")
    qdrant = QdrantClient(url=settings.qdrant_url)
    total_questions = 0

    for i, doc in enumerate(docs, 1):
        filename = doc["filename"]
        doc_id = doc["id"]

        print(f"\n  [{i}/{len(docs)}] {filename}")

        # Fetch chunks
        chunks = get_chunks(qdrant, doc_id)
        print(f"    Chunks: {len(chunks)}")

        # Create workbook with Chunks tab
        wb = Workbook()
        ws_chunk = wb.active
        ws_chunk.title = "Chunks"

        chunk_headers = ["Index", "Words", "Characters", "Page", "Section", "Text"]
        styled_header(ws_chunk, chunk_headers)

        for row_num, c in enumerate(chunks, 2):
            ws_chunk.cell(row=row_num, column=1, value=c["index"])
            ws_chunk.cell(row=row_num, column=2, value=c["words"])
            ws_chunk.cell(row=row_num, column=3, value=c["chars"])
            ws_chunk.cell(row=row_num, column=4, value=c["page"])
            ws_chunk.cell(row=row_num, column=5, value=_sanitize(c["section"] or ""))
            text_cell = ws_chunk.cell(row=row_num, column=6, value=_sanitize(c["text"]))
            text_cell.alignment = WRAP_ALIGN

        ws_chunk.column_dimensions["A"].width = 8
        ws_chunk.column_dimensions["B"].width = 8
        ws_chunk.column_dimensions["C"].width = 12
        ws_chunk.column_dimensions["D"].width = 8
        ws_chunk.column_dimensions["E"].width = 15
        ws_chunk.column_dimensions["F"].width = 100
        ws_chunk.freeze_panes = "A2"

        # Generate questions
        print("    Generating questions...", end=" ", flush=True)
        questions = generate_questions(chunks, filename, settings)
        print(f"{len(questions)} questions")
        total_questions += len(questions)

        # Questions tab
        ws_questions = wb.create_sheet("Questions")
        q_headers = ["#", "Question"]
        styled_header(ws_questions, q_headers)

        for row_num, q in enumerate(questions, 2):
            ws_questions.cell(row=row_num, column=1, value=row_num - 1)
            q_cell = ws_questions.cell(row=row_num, column=2, value=_sanitize(q))
            q_cell.alignment = WRAP_ALIGN

        ws_questions.column_dimensions["A"].width = 5
        ws_questions.column_dimensions["B"].width = 100
        ws_questions.freeze_panes = "A2"

        # Save per-document file
        safe_name = _safe_filename(filename)
        file_path = out_dir / f"{safe_name}_{today}.xlsx"
        wb.save(str(file_path))
        print(f"    Saved: {file_path.name}")

    print(f"\n{'=' * 60}")
    print(f"Exported to: {out_dir}/")
    print(f"  Files: {len(docs)}")
    print(f"  Total questions: {total_questions}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
