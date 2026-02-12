#!/usr/bin/env python3
"""Export document chunks and Ollama-generated questions to Excel — remote version.

Connects to a VE-RAG-System server via its REST API (no direct DB/Qdrant access needed).
Designed to run from a laptop targeting a remote Spark server over SSH tunnel or direct URL.

Output: data/chunks_MM_DD_YYYY/ (one .xlsx per document)

Usage:
    python scripts/export_chunks_remote.py                          # default: localhost:8502
    python scripts/export_chunks_remote.py --host 172.16.20.51      # target Spark directly
    python scripts/export_chunks_remote.py --host spark --port 8502  # SSH alias
    python scripts/export_chunks_remote.py -o custom_dir            # custom output directory
"""

import argparse
import re
from datetime import datetime
from pathlib import Path

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Regex to strip characters illegal in Excel XML (control chars except tab/newline/cr)
_ILLEGAL_XML_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]")


def _sanitize(text: str) -> str:
    """Remove characters that openpyxl/Excel cannot store in a cell."""
    return _ILLEGAL_XML_CHARS.sub("", text) if text else text


# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def get_api_client(base_url: str, email: str, password: str) -> tuple[httpx.Client, dict]:
    """Login and return (client, auth_headers)."""
    client = httpx.Client(base_url=base_url, timeout=30)
    r = client.post("/api/auth/login", json={"email": email, "password": password})
    r.raise_for_status()
    token = r.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}
    return client, headers


def get_ready_documents(client: httpx.Client, headers: dict) -> list[dict]:
    """Fetch all documents with status=ready via API."""
    docs = []
    page = 1
    while True:
        r = client.get(f"/api/documents?per_page=100&page={page}&status=ready", headers=headers)
        r.raise_for_status()
        data = r.json()
        docs.extend(data["documents"])
        if page >= data.get("total_pages", 1):
            break
        page += 1
    return docs


def get_chunks_via_api(client: httpx.Client, headers: dict, document_id: str) -> list[dict]:
    """Fetch chunks for a document via the search/experimental endpoint.

    Falls back to Qdrant direct if the API doesn't expose chunk listing.
    """
    # Try the document detail endpoint which may include chunks
    try:
        r = client.get(f"/api/documents/{document_id}/chunks", headers=headers, timeout=30)
        if r.status_code == 200:
            return r.json().get("chunks", [])
    except Exception:
        pass

    # Fallback: use the experimental search with a broad query scoped to this doc
    try:
        r = client.post(
            "/api/chat/search",
            json={"query": "*", "document_ids": [document_id], "top_k": 500},
            headers=headers,
            timeout=30,
        )
        if r.status_code == 200:
            results = r.json().get("results", [])
            chunks = []
            for i, result in enumerate(results):
                chunks.append(
                    {
                        "index": result.get("chunk_index", i),
                        "text": result.get("text", result.get("chunk_text", "")),
                        "words": len(result.get("text", result.get("chunk_text", "")).split()),
                        "chars": len(result.get("text", result.get("chunk_text", ""))),
                        "page": result.get("page_number"),
                        "section": result.get("section"),
                    }
                )
            return sorted(chunks, key=lambda c: c["index"])
    except Exception:
        pass

    return []


def get_chunks_via_qdrant(qdrant_url: str, document_id: str) -> list[dict]:
    """Fetch chunks directly from Qdrant (when API doesn't expose them)."""
    try:
        from qdrant_client import QdrantClient
        from qdrant_client.models import FieldCondition, Filter, MatchValue

        qdrant = QdrantClient(url=qdrant_url)
        points, _ = qdrant.scroll(
            collection_name="documents",
            scroll_filter=Filter(
                must=[FieldCondition(key="document_id", match=MatchValue(value=document_id))]
            ),
            with_payload=True,
            with_vectors=False,
            limit=1000,
        )

        chunks = []
        for p in points:
            payload = p.payload or {}
            chunks.append(
                {
                    "index": payload.get("chunk_index", 0),
                    "text": payload.get("chunk_text", ""),
                    "words": len(payload.get("chunk_text", "").split()),
                    "chars": len(payload.get("chunk_text", "")),
                    "page": payload.get("page_number"),
                    "section": payload.get("section"),
                }
            )
        return sorted(chunks, key=lambda c: c["index"])
    except Exception as e:
        print(f"    WARNING: Qdrant access failed: {e}")
        return []


def generate_questions_remote(
    ollama_url: str, model: str, chunks: list[dict], doc_name: str
) -> list[str]:
    """Ask Ollama on the remote server for recommended questions."""
    if not chunks:
        return []

    sample_size = min(12, len(chunks))
    step = max(1, len(chunks) // sample_size)
    sampled = [chunks[i] for i in range(0, len(chunks), step)][:sample_size]

    total_words = sum(c["words"] for c in chunks)
    content_summary = "\n\n".join(f"[Chunk {c['index']}]: {c['text'][:300]}" for c in sampled)

    if len(chunks) <= 3:
        num_questions = 3
    elif len(chunks) <= 10:
        num_questions = 5
    elif len(chunks) <= 50:
        num_questions = 10
    else:
        num_questions = 15

    prompt = f"""You are analyzing a document that has been chunked for a RAG system.
Document: {doc_name}
Total chunks: {len(chunks)}
Total words: {total_words:,}

Here are representative samples from the document:

{content_summary}

Based on this content, suggest exactly {num_questions} specific questions that a user could ask this RAG system.
Focus on questions that:
1. Can be answered from the document content
2. Cover different sections/topics in the document
3. Range from simple factual to analytical
4. Would be useful for someone who needs information from this document

Format: Number each question on its own line (1. question text). No other text."""

    try:
        with httpx.Client(timeout=300) as client:
            response = client.post(
                f"{ollama_url}/api/generate",
                json={
                    "model": model,
                    "prompt": prompt,
                    "stream": False,
                    "options": {"temperature": 0.3},
                },
            )
            response.raise_for_status()
            answer = response.json().get("response", "").strip()

            questions = []
            for line in answer.split("\n"):
                line = line.strip()
                if line and line[0].isdigit():
                    q = line.lstrip("0123456789").lstrip(".)- ").strip()
                    if q:
                        questions.append(q)
            return questions

    except httpx.TimeoutException:
        print(f"    WARNING: Ollama timed out for '{doc_name}'")
        return [f"[Ollama timed out — try: ollama pull {model}]"]
    except httpx.ConnectError:
        print(f"    WARNING: Cannot connect to Ollama at {ollama_url}")
        return ["[Ollama not reachable]"]
    except Exception as e:
        print(f"    WARNING: Ollama error for '{doc_name}': {e}")
        return [f"[Error: {e}]"]


def styled_header(ws, headers: list[str], row: int = 1) -> None:
    """Write a styled header row."""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _safe_filename(name: str) -> str:
    """Convert a document name into a filesystem-safe filename (no extension)."""
    stem = Path(name).stem
    safe = re.sub(r'[\\/*?:"<>|]', "_", stem)
    safe = re.sub(r"[\s_]+", "_", safe).strip("_")
    return safe[:120]


def main():
    parser = argparse.ArgumentParser(
        description="Export per-document chunk + question Excel files (remote)"
    )
    parser.add_argument("--host", default="172.16.20.51", help="Server host (default: Spark)")
    parser.add_argument("--port", default="8502", help="Server port (default: 8502)")
    parser.add_argument("--email", default="admin@test.com", help="Login email")
    parser.add_argument("--password", default="npassword2002!", help="Login password")
    parser.add_argument("--ollama-url", default=None, help="Ollama URL (default: same host:11434)")
    parser.add_argument("--qdrant-url", default=None, help="Qdrant URL (default: same host:6333)")
    parser.add_argument("--model", default="qwen3:8b", help="Ollama model for questions")
    parser.add_argument(
        "-o", "--output", help="Output directory (default: data/chunks_MM_DD_YYYY/)"
    )
    parser.add_argument(
        "--no-questions", action="store_true", help="Skip Ollama question generation"
    )
    args = parser.parse_args()

    base_url = f"http://{args.host}:{args.port}"
    ollama_url = args.ollama_url or f"http://{args.host}:11434"
    qdrant_url = args.qdrant_url or f"http://{args.host}:6333"
    today = datetime.now().strftime("%m_%d_%Y")
    out_dir = Path(args.output) if args.output else Path("data") / f"chunks_{today}"
    out_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("RAG Document & Chunk Export (remote)")
    print(f"  Server:  {base_url}")
    print(f"  Ollama:  {ollama_url} (model: {args.model})")
    print(f"  Qdrant:  {qdrant_url}")
    print(f"  Output:  {out_dir}/")
    print("=" * 60)

    # Login
    print("\nConnecting...", end=" ", flush=True)
    client, headers = get_api_client(base_url, args.email, args.password)
    print("OK")

    # Fetch documents
    print("Fetching documents...", end=" ", flush=True)
    docs = get_ready_documents(client, headers)
    print(f"{len(docs)} ready documents")

    if not docs:
        print("No ready documents to export.")
        return

    total_questions = 0

    for i, doc in enumerate(docs, 1):
        filename = doc.get("original_filename") or doc.get("filename", "unknown")
        doc_id = doc["id"]

        print(f"\n  [{i}/{len(docs)}] {filename}")

        # Fetch chunks (try API first, then Qdrant direct)
        chunks = get_chunks_via_api(client, headers, doc_id)
        if not chunks:
            chunks = get_chunks_via_qdrant(qdrant_url, doc_id)
        print(f"    Chunks: {len(chunks)}")

        if not chunks:
            print("    SKIP: no chunks found")
            continue

        # Create workbook
        wb = Workbook()
        ws_chunk = wb.active
        ws_chunk.title = "Chunks"

        chunk_headers = ["Index", "Words", "Characters", "Page", "Section", "Text"]
        styled_header(ws_chunk, chunk_headers)

        for row_num, c in enumerate(chunks, 2):
            ws_chunk.cell(row=row_num, column=1, value=c["index"])
            ws_chunk.cell(row=row_num, column=2, value=c["words"])
            ws_chunk.cell(row=row_num, column=3, value=c["chars"])
            ws_chunk.cell(row=row_num, column=4, value=c["page"])
            ws_chunk.cell(row=row_num, column=5, value=_sanitize(c.get("section") or ""))
            text_cell = ws_chunk.cell(row=row_num, column=6, value=_sanitize(c["text"]))
            text_cell.alignment = WRAP_ALIGN

        ws_chunk.column_dimensions["A"].width = 8
        ws_chunk.column_dimensions["B"].width = 8
        ws_chunk.column_dimensions["C"].width = 12
        ws_chunk.column_dimensions["D"].width = 8
        ws_chunk.column_dimensions["E"].width = 15
        ws_chunk.column_dimensions["F"].width = 100
        ws_chunk.freeze_panes = "A2"

        # Generate questions
        if not args.no_questions:
            print("    Generating questions...", end=" ", flush=True)
            questions = generate_questions_remote(ollama_url, args.model, chunks, filename)
            print(f"{len(questions)} questions")
            total_questions += len(questions)

            ws_questions = wb.create_sheet("Questions")
            q_headers = ["#", "Question"]
            styled_header(ws_questions, q_headers)

            for row_num, q in enumerate(questions, 2):
                ws_questions.cell(row=row_num, column=1, value=row_num - 1)
                q_cell = ws_questions.cell(row=row_num, column=2, value=_sanitize(q))
                q_cell.alignment = WRAP_ALIGN

            ws_questions.column_dimensions["A"].width = 5
            ws_questions.column_dimensions["B"].width = 100
            ws_questions.freeze_panes = "A2"

        # Save
        safe_name = _safe_filename(filename)
        file_path = out_dir / f"{safe_name}_{today}.xlsx"
        wb.save(str(file_path))
        print(f"    Saved: {file_path.name}")

    client.close()

    print(f"\n{'=' * 60}")
    print(f"Exported to: {out_dir}/")
    print(f"  Files: {len(docs)}")
    if not args.no_questions:
        print(f"  Total questions: {total_questions}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
