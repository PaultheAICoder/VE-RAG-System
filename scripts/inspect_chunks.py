#!/usr/bin/env python3
"""Inspect vector chunks for a document.

Shows what data was extracted and optionally asks Ollama for recommended questions.

Usage:
    # Summary view (default) — first 200 chars per chunk
    python scripts/inspect_chunks.py DOCUMENT_ID

    # Detail view — full chunk text
    python scripts/inspect_chunks.py DOCUMENT_ID --detail

    # Ask Ollama for recommended questions
    python scripts/inspect_chunks.py DOCUMENT_ID --recommend

    # Output to Excel (includes all chunks + recommendations sheet)
    python scripts/inspect_chunks.py DOCUMENT_ID --output chunks.xlsx
    python scripts/inspect_chunks.py DOCUMENT_ID --output chunks.xlsx --recommend

    # Limit chunks shown (console only, Excel always includes all)
    python scripts/inspect_chunks.py DOCUMENT_ID --detail --limit 10
"""

import argparse
import sys
from pathlib import Path

import httpx
from qdrant_client import QdrantClient
from qdrant_client.models import FieldCondition, Filter, MatchValue

# Add project root to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from ai_ready_rag.config import get_settings


def get_chunks(qdrant_url: str, document_id: str) -> list[dict]:
    """Fetch all chunks for a document from Qdrant."""
    client = QdrantClient(url=qdrant_url)

    points, _ = client.scroll(
        collection_name="documents",
        scroll_filter=Filter(
            must=[FieldCondition(key="document_id", match=MatchValue(value=document_id))]
        ),
        with_payload=True,
        with_vectors=False,
        limit=1000,
    )

    chunks = []
    for p in points:
        payload = p.payload or {}
        chunks.append(
            {
                "index": payload.get("chunk_index", 0),
                "text": payload.get("chunk_text", ""),
                "page": payload.get("page_number"),
                "section": payload.get("section"),
                "tags": payload.get("tags", []),
                "document_name": payload.get("document_name", ""),
            }
        )

    return sorted(chunks, key=lambda c: c["index"])


def get_recommendations(chunks: list[dict], settings) -> str | None:
    """Ask Ollama to suggest questions based on the document content.

    Returns the recommendation text, or None on failure.
    """
    sample_size = min(12, len(chunks))
    step = max(1, len(chunks) // sample_size)
    sampled = [chunks[i] for i in range(0, len(chunks), step)][:sample_size]

    doc_name = chunks[0]["document_name"] if chunks else "Unknown"
    content_summary = "\n\n".join(f"[Chunk {c['index']}]: {c['text'][:300]}" for c in sampled)

    prompt = f"""You are analyzing a document that has been chunked for a RAG system.
Document: {doc_name}
Total chunks: {len(chunks)}
Total words: {sum(len(c["text"].split()) for c in chunks):,}

Here are representative samples from the document:

{content_summary}

Based on this content, suggest 10 specific questions that a user could ask this RAG system.
Focus on questions that:
1. Can be answered from the document content
2. Cover different sections/topics in the document
3. Range from simple factual to analytical
4. Would be useful for someone who needs information from this document

Format: Number each question on its own line. No other text."""

    ollama_url = settings.ollama_base_url
    model = settings.chat_model

    print(f"Generating recommendations via {model}...")
    print(f"  Sampling {len(sampled)} of {len(chunks)} chunks...")

    try:
        with httpx.Client(timeout=300) as client:
            response = client.post(
                f"{ollama_url}/api/generate",
                json={
                    "model": model,
                    "prompt": prompt,
                    "stream": False,
                    "options": {"temperature": 0.3},
                },
            )
            response.raise_for_status()
            return response.json().get("response", "").strip()
    except httpx.TimeoutException:
        print("ERROR: Ollama request timed out. Is the model loaded?")
        print(f"  Try: ollama pull {model}")
    except httpx.ConnectError:
        print(f"ERROR: Cannot connect to Ollama at {ollama_url}")
        print("  Is Ollama running? Try: ollama serve")
    except Exception as e:
        print(f"ERROR: {e}")
    return None


def write_excel(
    chunks: list[dict],
    document_id: str,
    output_path: str,
    recommendations: str | None = None,
) -> None:
    """Write chunks to an Excel workbook with summary + detail sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    doc_name = chunks[0]["document_name"] if chunks else "Unknown"
    tags = chunks[0]["tags"] if chunks else []
    total_words = sum(len(c["text"].split()) for c in chunks)
    total_chars = sum(len(c["text"]) for c in chunks)
    sizes = [len(c["text"]) for c in chunks]

    summary_data = [
        ("Document", doc_name),
        ("Document ID", document_id),
        ("Tags", ", ".join(tags) if tags else "none"),
        ("Total Chunks", len(chunks)),
        ("Total Words", total_words),
        ("Total Characters", total_chars),
        ("Avg Words/Chunk", total_words // len(chunks) if chunks else 0),
        ("Min Chunk Size", min(sizes) if sizes else 0),
        ("Max Chunk Size", max(sizes) if sizes else 0),
        ("Median Chunk Size", sorted(sizes)[len(sizes) // 2] if sizes else 0),
    ]

    label_font = Font(bold=True)
    for row_num, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_num, column=1, value=label).font = label_font
        ws_summary.cell(row=row_num, column=2, value=value)

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 50

    # --- Chunks sheet ---
    ws_chunks = wb.create_sheet("Chunks")

    chunk_headers = ["Index", "Words", "Characters", "Page", "Section", "Text"]
    for col_num, header in enumerate(chunk_headers, 1):
        cell = ws_chunks.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_num, c in enumerate(chunks, 2):
        words = len(c["text"].split())
        ws_chunks.cell(row=row_num, column=1, value=c["index"])
        ws_chunks.cell(row=row_num, column=2, value=words)
        ws_chunks.cell(row=row_num, column=3, value=len(c["text"]))
        ws_chunks.cell(row=row_num, column=4, value=c["page"])
        ws_chunks.cell(row=row_num, column=5, value=c["section"])
        text_cell = ws_chunks.cell(row=row_num, column=6, value=c["text"])
        text_cell.alignment = wrap_alignment

    # Column widths
    ws_chunks.column_dimensions["A"].width = 8
    ws_chunks.column_dimensions["B"].width = 8
    ws_chunks.column_dimensions["C"].width = 12
    ws_chunks.column_dimensions["D"].width = 8
    ws_chunks.column_dimensions["E"].width = 15
    ws_chunks.column_dimensions["F"].width = 100
    ws_chunks.freeze_panes = "A2"

    # --- Recommendations sheet (if available) ---
    if recommendations:
        ws_rec = wb.create_sheet("Recommended Questions")

        cell = ws_rec.cell(row=1, column=1, value="Recommended Questions")
        cell.font = Font(bold=True, size=14)

        ws_rec.cell(row=2, column=1, value=f"Document: {doc_name}")
        ws_rec.cell(row=3, column=1, value=f"Generated by: {get_settings().chat_model}")
        ws_rec.cell(row=4, column=1, value="")

        # Parse numbered questions
        questions = []
        for line in recommendations.split("\n"):
            line = line.strip()
            if line and line[0].isdigit():
                # Strip leading number and punctuation
                q = line.lstrip("0123456789").lstrip(".)- ").strip()
                if q:
                    questions.append(q)

        if questions:
            header_cell = ws_rec.cell(row=5, column=1, value="#")
            header_cell.font = header_font
            header_cell.fill = header_fill
            q_header = ws_rec.cell(row=5, column=2, value="Question")
            q_header.font = header_font
            q_header.fill = header_fill

            for i, q in enumerate(questions, 1):
                ws_rec.cell(row=5 + i, column=1, value=i)
                cell = ws_rec.cell(row=5 + i, column=2, value=q)
                cell.alignment = wrap_alignment
        else:
            # Couldn't parse — dump raw text
            ws_rec.cell(row=5, column=1, value=recommendations).alignment = wrap_alignment

        ws_rec.column_dimensions["A"].width = 5
        ws_rec.column_dimensions["B"].width = 100

    wb.save(output_path)


def print_summary(chunks: list[dict], preview_chars: int = 200) -> None:
    """Print summary view — chunk index, word count, preview."""
    total_words = sum(len(c["text"].split()) for c in chunks)
    total_chars = sum(len(c["text"]) for c in chunks)
    doc_name = chunks[0]["document_name"] if chunks else "Unknown"
    tags = chunks[0]["tags"] if chunks else []

    print(f"Document:    {doc_name}")
    print(f"Tags:        {', '.join(tags) if tags else 'none'}")
    print(f"Chunks:      {len(chunks)}")
    print(f"Total words: {total_words:,}")
    print(f"Total chars: {total_chars:,}")
    print(f"Avg words/chunk: {total_words // len(chunks) if chunks else 0}")
    print()

    sizes = [len(c["text"]) for c in chunks]
    print(
        f"Chunk sizes: min={min(sizes)}, max={max(sizes)}, median={sorted(sizes)[len(sizes) // 2]}"
    )
    print("=" * 80)

    for c in chunks:
        words = len(c["text"].split())
        page_str = f"p.{c['page']}" if c["page"] else ""
        section_str = c["section"] or ""
        meta = " | ".join(filter(None, [page_str, section_str]))
        meta_display = f" [{meta}]" if meta else ""

        preview = c["text"][:preview_chars].replace("\n", " ")
        if len(c["text"]) > preview_chars:
            preview += "..."

        print(f"\n  [{c['index']:3d}] ({words:4d} words){meta_display}")
        print(f"        {preview}")


def print_detail(chunks: list[dict], limit: int | None = None) -> None:
    """Print full chunk text."""
    display = chunks[:limit] if limit else chunks

    for c in display:
        words = len(c["text"].split())
        page_str = f"page={c['page']}" if c["page"] else ""
        section_str = f"section={c['section']}" if c["section"] else ""
        meta = " | ".join(filter(None, [page_str, section_str]))

        print(f"\n{'=' * 80}")
        print(f"Chunk {c['index']} | {words} words | {len(c['text'])} chars | {meta}")
        print("-" * 80)
        print(c["text"])

    if limit and limit < len(chunks):
        print(f"\n... showing {limit} of {len(chunks)} chunks (use --limit to adjust)")


def print_recommendations(recommendations: str) -> None:
    """Print recommendations to stdout."""
    print("\n" + "=" * 80)
    print("RECOMMENDED QUESTIONS (via Ollama)")
    print("=" * 80)
    print(recommendations)


def main():
    parser = argparse.ArgumentParser(
        description="Inspect vector chunks for a document",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Tip: Use scripts/list_documents.py to find document IDs",
    )
    parser.add_argument("document_id", help="Document ID (UUID)")
    parser.add_argument("--detail", action="store_true", help="Show full chunk text")
    parser.add_argument(
        "--recommend", action="store_true", help="Ask Ollama for recommended questions"
    )
    parser.add_argument("--limit", type=int, help="Limit number of chunks shown (detail mode)")
    parser.add_argument(
        "--preview",
        type=int,
        default=200,
        help="Preview chars per chunk in summary (default: 200)",
    )
    parser.add_argument("-o", "--output", help="Write to Excel file (e.g. chunks.xlsx)")
    args = parser.parse_args()

    settings = get_settings()
    chunks = get_chunks(settings.qdrant_url, args.document_id)

    if not chunks:
        print(f"No chunks found for document {args.document_id}")
        print("Check the ID with: python scripts/list_documents.py --format ids")
        sys.exit(1)

    # Get recommendations if requested (needed for both console and Excel)
    recommendations = None
    if args.recommend:
        recommendations = get_recommendations(chunks, settings)

    if args.output:
        write_excel(chunks, args.document_id, args.output, recommendations)
        doc_name = chunks[0]["document_name"] if chunks else "Unknown"
        print(f"Wrote {len(chunks)} chunks from '{doc_name}' to {args.output}")
        if recommendations:
            print("  Includes 'Recommended Questions' sheet")
    else:
        # Console output
        if args.detail:
            print_detail(chunks, limit=args.limit)
        else:
            print_summary(chunks, preview_chars=args.preview)

        if recommendations:
            print_recommendations(recommendations)


if __name__ == "__main__":
    main()
