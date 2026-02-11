#!/usr/bin/env python3
"""List all documents stored in the vector database.

Usage:
    python scripts/list_documents.py
    python scripts/list_documents.py --status ready
    python scripts/list_documents.py --tag hr
    python scripts/list_documents.py --output documents.xlsx
"""

import argparse
import sys
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from ai_ready_rag.db.database import SessionLocal, init_db
from ai_ready_rag.db.models import Document, Tag, document_tags


def build_rows(docs) -> list[dict]:
    """Convert Document ORM objects to flat dicts."""
    rows = []
    for doc in docs:
        rows.append(
            {
                "id": doc.id,
                "filename": doc.original_filename or doc.filename,
                "file_type": doc.file_type,
                "file_size": doc.file_size,
                "status": doc.status,
                "chunk_count": doc.chunk_count,
                "word_count": doc.word_count,
                "processing_time_ms": doc.processing_time_ms,
                "error_message": doc.error_message,
                "tags": ", ".join(t.name for t in doc.tags) if doc.tags else "",
                "uploaded_by": doc.uploaded_by,
                "uploaded_at": str(doc.uploaded_at) if doc.uploaded_at else "",
            }
        )
    return rows


def write_excel(rows: list[dict], output_path: str) -> None:
    """Write rows to an Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"

    if not rows:
        ws.append(["No documents found"])
        wb.save(output_path)
        return

    # Header
    headers = list(rows[0].keys())
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_num, _header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row in rows:
        ws.append(list(row.values()))

    # Auto-width columns
    for col_num in range(1, len(headers) + 1):
        letter = get_column_letter(col_num)
        max_len = max(
            len(str(ws.cell(row=r, column=col_num).value or "")) for r in range(1, len(rows) + 2)
        )
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)


def print_table(rows: list[dict]) -> None:
    """Print rows as a formatted table to stdout."""
    print(
        f"{'ID':<38} {'Filename':<40} {'Type':<5} {'Status':<10} "
        f"{'Chunks':>6} {'Words':>7} {'Tags'}"
    )
    print("-" * 140)

    for row in rows:
        filename = row["filename"]
        if len(filename) > 38:
            filename = filename[:35] + "..."
        chunks = str(row["chunk_count"]) if row["chunk_count"] else "-"
        words = str(row["word_count"]) if row["word_count"] else "-"

        print(
            f"{row['id']:<38} {filename:<40} {row['file_type']:<5} {row['status']:<10} "
            f"{chunks:>6} {words:>7} {row['tags']}"
        )

    print(f"\nTotal: {len(rows)} documents")

    statuses = {}
    for row in rows:
        statuses[row["status"]] = statuses.get(row["status"], 0) + 1
    status_str = ", ".join(f"{s}: {c}" for s, c in sorted(statuses.items()))
    print(f"Status: {status_str}")


def main():
    parser = argparse.ArgumentParser(description="List documents in the RAG system")
    parser.add_argument("--status", help="Filter by status (pending, processing, ready, failed)")
    parser.add_argument("--tag", help="Filter by tag name")
    parser.add_argument("--format", choices=["table", "ids"], default="table", help="Output format")
    parser.add_argument("-o", "--output", help="Write to Excel file (e.g. documents.xlsx)")
    args = parser.parse_args()

    init_db()
    db = SessionLocal()

    try:
        query = db.query(Document).order_by(Document.uploaded_at.desc())

        if args.status:
            query = query.filter(Document.status == args.status)

        if args.tag:
            query = query.join(document_tags).join(Tag).filter(Tag.name == args.tag)

        docs = query.all()

        if not docs:
            print("No documents found.")
            return

        rows = build_rows(docs)

        if args.output:
            write_excel(rows, args.output)
            print(f"Wrote {len(rows)} documents to {args.output}")
        elif args.format == "ids":
            for row in rows:
                print(f"{row['id']}  {row['filename']}")
        else:
            print_table(rows)

    finally:
        db.close()


if __name__ == "__main__":
    main()
