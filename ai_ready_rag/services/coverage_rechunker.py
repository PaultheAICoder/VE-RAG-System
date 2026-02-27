"""Coverage Summary Excel rechunker.

Replaces Docling's low-quality xlsx chunks with clean, section-based
natural-text chunks for insurance Coverage Summary spreadsheets.

Docling's xlsx handler produces repetitive "key, ENTITY = value" chunks
that dilute embeddings. This rechunker reads the raw xlsx via openpyxl,
groups rows by insurance section headers, and formats each section as
natural text with clear labels.
"""

import logging
import re
from pathlib import Path

from sqlalchemy.orm import Session

from ai_ready_rag.config import Settings
from ai_ready_rag.db.models import Document

logger = logging.getLogger(__name__)

# Insurance section header keywords (case-insensitive)
SECTION_KEYWORDS = [
    "general liability",
    "property",
    "workers compensation",
    "workers comp",
    "directors & officers",
    "directors and officers",
    "d&o",
    "crime",
    "auto",
    "automobile",
    "umbrella",
    "excess",
    "professional liability",
    "errors & omissions",
    "e&o",
    "cyber",
    "inland marine",
    "employment practices",
    "epli",
    "fiduciary",
    "pollution",
    "environmental",
    "flood",
    "earthquake",
    "builder",
    "boiler",
    "equipment breakdown",
]


def is_coverage_summary(document: Document) -> bool:
    """Detect whether a document is a Coverage Summary xlsx.

    Checks filename for "coverage" or "summary" keywords,
    or if the document has an 'insurance' tag.
    """
    fname = (document.original_filename or "").lower()
    if "coverage" in fname or "summary" in fname:
        return True

    tag_names = {tag.name.lower() for tag in document.tags}
    return "insurance" in tag_names


def _is_section_header(row_values: list, bold_flags: list[bool] | None = None) -> str | None:
    """Check if a row is a section header.

    Returns the section name if it matches, None otherwise.
    """
    # Combine all non-empty cell values into one string
    text = " ".join(str(v).strip() for v in row_values if v is not None and str(v).strip())
    if not text:
        return None

    # Only match if the row is primarily a section header — not a data row
    # that happens to contain a keyword (e.g. "Business Personal Property: $500K")
    text_lower = text.lower().strip()
    for kw in SECTION_KEYWORDS:
        if text_lower == kw or text_lower.startswith(kw + " ") or text_lower.endswith(" " + kw):
            return text
        # Also match "Property Insurance", "General Liability Coverage", etc.
        if kw in text_lower and len(text) < 60 and not re.search(r"[\$%]", text):
            return text

    # Check if the row is bold (common for section headers)
    if (
        bold_flags
        and any(bold_flags)
        and 3 < len(text) < 80
        and not re.search(r"[\$%\d]{3,}", text)
    ):
        return text

    return None


def _extract_entity_name(ws) -> str:
    """Try to extract the entity/insured name from the first few rows."""
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and len(cell) > 5:
                val = cell.strip()
                # Skip generic headers
                if val.lower() in ("coverage summary", "insurance summary", "policy summary"):
                    continue
                return val
    return "Coverage Summary"


def _extract_year_range(ws) -> str:
    """Try to extract the policy year range from the first few rows."""
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                # Look for year ranges like "2025-2026" or "2025/2026"
                match = re.search(r"20\d{2}[-/]20\d{2}", cell)
                if match:
                    return match.group()
    return ""


def parse_coverage_sections(file_path: Path) -> list[dict]:
    """Parse a Coverage Summary xlsx into section-based chunks.

    Args:
        file_path: Path to the xlsx file.

    Returns:
        List of dicts with 'text' and 'section' keys.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    ws = wb.active
    if ws is None:
        logger.warning("No active worksheet found in %s", file_path)
        return []

    entity_name = _extract_entity_name(ws)
    year_range = _extract_year_range(ws)
    header_line = f"{entity_name}"
    if year_range:
        header_line += f" - Coverage Summary {year_range}"

    # Scan all rows, grouping into sections
    sections: list[dict] = []
    current_section: str | None = None
    current_rows: list[str] = []

    for row in ws.iter_rows(min_row=1):
        values = [cell.value for cell in row]
        bold_flags = [getattr(cell.font, "bold", False) if cell.font else False for cell in row]

        section_name = _is_section_header(values, bold_flags)
        if section_name:
            # Flush previous section
            if current_section and current_rows:
                sections.append(
                    {
                        "section": current_section,
                        "rows": current_rows,
                    }
                )
            current_section = section_name
            current_rows = []
            continue

        # Skip empty rows
        non_empty = [v for v in values if v is not None and str(v).strip()]
        if not non_empty:
            continue

        # Format row as key-value pairs
        # Common patterns: [label, value] or [label, "", value]
        cleaned = [str(v).strip() for v in values if v is not None and str(v).strip()]
        if len(cleaned) == 1:
            current_rows.append(cleaned[0])
        elif len(cleaned) == 2:
            current_rows.append(f"{cleaned[0]}: {cleaned[1]}")
        else:
            # Join with appropriate separator
            current_rows.append(": ".join(cleaned[:2]) + " " + " ".join(cleaned[2:]))

    # Flush last section
    if current_section and current_rows:
        sections.append(
            {
                "section": current_section,
                "rows": current_rows,
            }
        )

    # If no sections detected, create one chunk from all rows
    if not sections and current_rows:
        sections.append(
            {
                "section": "General",
                "rows": current_rows,
            }
        )

    # Convert sections to chunk text
    chunks = []
    for section in sections:
        lines = [header_line, "", f"{section['section']}", ""]
        lines.extend(section["rows"])
        chunk_text = "\n".join(lines)

        # Skip very short sections (< 3 data rows)
        if len(section["rows"]) < 2:
            continue

        chunks.append(
            {
                "text": chunk_text,
                "section": section["section"],
            }
        )

    return chunks


async def rechunk_coverage_summary(
    document: Document,
    db: Session,
    settings: Settings,
) -> int:
    """Rechunk a Coverage Summary xlsx with clean section-based chunks.

    Replaces existing vector store chunks for this document with
    natural-text section chunks parsed from the raw xlsx.

    Args:
        document: Document record (must be xlsx, status=ready).
        db: Database session.
        settings: Application settings.

    Returns:
        Number of new chunks written (0 if skipped or failed).
    """
    file_path = Path(settings.upload_dir) / document.stored_filename

    if not file_path.exists():
        logger.warning(
            "coverage_rechunk.file_not_found: %s for doc %s",
            file_path,
            document.id,
        )
        return 0

    # Parse sections from raw xlsx
    try:
        section_chunks = parse_coverage_sections(file_path)
    except Exception as e:
        logger.error(
            "coverage_rechunk.parse_failed: doc=%s error=%s",
            document.id,
            e,
        )
        return 0

    if not section_chunks:
        logger.info(
            "coverage_rechunk.no_sections: doc=%s — keeping original chunks",
            document.id,
        )
        return 0

    # Replace chunks via vector service (handles delete + embed + upsert)
    from ai_ready_rag.services.factory import get_vector_service

    vector_service = get_vector_service(settings)

    tag_names = [tag.name for tag in document.tags]
    chunk_texts = [c["text"] for c in section_chunks]
    chunk_metadata = [{"section": c["section"]} for c in section_chunks]

    try:
        result = await vector_service.add_document(
            document_id=document.id,
            document_name=document.original_filename,
            chunks=chunk_texts,
            tags=tag_names,
            uploaded_by=document.uploaded_by,
            chunk_metadata=chunk_metadata,
        )

        new_count = result.chunk_count
        logger.info(
            "coverage_rechunk.success: doc=%s sections=%d chunks=%d",
            document.id,
            len(section_chunks),
            new_count,
        )

        # Update document chunk count
        document.chunk_count = new_count
        db.commit()

        return new_count

    except Exception as e:
        logger.error(
            "coverage_rechunk.upsert_failed: doc=%s error=%s",
            document.id,
            e,
        )
        return 0
